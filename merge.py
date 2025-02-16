#!/usr/bin/env python
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

DATA_DIR = Path('data/filtered')
BOOK = 'bom.xlsx'

# Open or create workbook
if Path(BOOK).exists():
    book = load_workbook(BOOK)
else:
    book = Workbook()
    # Remove the default sheet created by openpyxl
    default_sheet = book.active
    book.remove(default_sheet)

# Process CSVs and edit "link" column
for csv_file in DATA_DIR.iterdir():
    if csv_file.suffix == '.csv':  # Ensure only CSV files are processed
        df = pd.read_csv(csv_file)
        df['link'] = df.pop('link').apply(lambda x: f'=HYPERLINK("{x}")')

        # Add or overwrite a worksheet
        sheet_name = csv_file.stem
        if sheet_name in book.sheetnames:
            book.remove(book[sheet_name])
        sheet = book.create_sheet(title=sheet_name)

        # Write DataFrame to the worksheet
        for row in dataframe_to_rows(df, index=False, header=True):
            sheet.append(row)

# Save workbook
book.save(BOOK)
